# -*- coding: utf-8 -*-
"""
Created on Mon Dec  7 00:29:20 2020

@author: DELL
"""


'''
"code":
"contentId":
"des":小概括******************2
"imgUrl":图片url
"keyword":文章标签*************4
"listResult":*****************
"pubtime":报道时间*************6
"sitename":网站名************7
"title":新闻标题**************8
"url":具体新闻的网址************9
""
'''

import requests
import json
import jsonpath
from bs4 import BeautifulSoup
import openpyxl 
import time

#12.8-6.20
#
#441 6.20
#
def getOnePageData(url):
    r = requests.get(url)
    html = json.loads(r.text)
    onePageData = jsonpath.jsonpath(html, '$..results')[0]
    return onePageData


def saveOnePageData(filePath, onePageData):
    for j in range(10):
        try:
            worksheet.append([onePageData[j]["pubtime"], onePageData[j]["sitename"], onePageData[j]["title"], onePageData[j]["url"]])
        except IndexError as e:
            print(e)    
    workbook.save(filePath)


def getContent(url0):
    r0 = requests.get(url0)
    r0.encoding = "utf8"
    soup = BeautifulSoup(r0.text, 'lxml')
    content = soup.find_all('p')   
    for i in range(0,len(content)):
        content[i] = content[i].get_text()
    return "".join(content)


def saveContent():
    for k in range(1, len(list(worksheet.columns)[3])):
        if list(worksheet.columns)[1][k].value[:4] == "测试站点":
            continue
        if list(worksheet.columns)[2][k].value[:3] == "陈一新":
            continue
        url0 = list(worksheet.columns)[3][k].value
        worksheet.cell(k+1, 5, str(getContent(url0)))
        workbook.save(filePath)
        print(k)
        


if __name__ == "__main__":
    #创建文件
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    #worksheet.append(['code', 'contentId', 'des', 'imgUrl', 'keyword', 'listResult', 'pubtime', 'sitename', 'title', 'url'])
    worksheet.append(['pubtime', 'sitename', 'title', 'url'])
    filePath = r'E:\lhy\sophomore First\Fundamentals of Data Science\BIGBIGBIG\xinhua.xlsx'
    for i in range(441, 1000):
        url = 'http://so.news.cn/getNews?keyword=%E7%96%AB&curPage='+str(i)+'&sortField=0&searchFields=1&lang=cn'
        onePageData = getOnePageData(url)
        if onePageData == None:
            time.sleep(10)
            onePageData = getOnePageData(url)
        print("finish",i)
        saveOnePageData(filePath, onePageData)
    
    workbook = openpyxl.load_workbook(filePath)
    worksheet = workbook.worksheets[0]
    worksheet.cell(1, 5, "content")
    workbook.save(filePath)
    saveContent()